# ============================================================
# excel_writer.py
# ============================================================
# Convert generated CSV sign-in sheets into formatted Excel files.
# ============================================================

# === IMPORTS ===

import csv
import logging
from pathlib import Path
from typing import Final

from datafun_toolkit.logger import get_logger, log_path
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# === CONFIGURE LOGGER ===

LOG: logging.Logger = get_logger("P6", level="DEBUG")

# === STYLE CONSTANTS ===

THIN_BLACK_SIDE: Final[Side] = Side(border_style="thin", color="FF000000")
MEDIUM_BLACK_SIDE: Final[Side] = Side(border_style="medium", color="FF000000")

THIN_BORDER: Final[Border] = Border(
    left=THIN_BLACK_SIDE,
    right=THIN_BLACK_SIDE,
    top=THIN_BLACK_SIDE,
    bottom=THIN_BLACK_SIDE,
)

HEADER_BORDER: Final[Border] = Border(
    left=MEDIUM_BLACK_SIDE,
    right=MEDIUM_BLACK_SIDE,
    top=MEDIUM_BLACK_SIDE,
    bottom=MEDIUM_BLACK_SIDE,
)

HEADER_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    start_color="FFD9D9D9",
    end_color="FFD9D9D9",
)

HEADER_FONT: Final[Font] = Font(bold=True)

# === PRINT CONSTANTS ===

LETTER_PAPER_SIZE: Final[int] = 1

DEFAULT_LEFT_MARGIN: Final[float] = 0.25
DEFAULT_RIGHT_MARGIN: Final[float] = 0.25
DEFAULT_TOP_MARGIN: Final[float] = 1.0
DEFAULT_BOTTOM_MARGIN: Final[float] = 0.75
DEFAULT_HEADER_MARGIN: Final[float] = 0.3
DEFAULT_FOOTER_MARGIN: Final[float] = 0.3


# === PUBLIC FUNCTIONS ===


def convert_csv_to_formatted_xlsx(
    csv_path: Path,
    xlsx_path: Path | None = None,
    *,
    title_prefix: str = "Members Signin",
    landscape: bool = True,
    repeat_header_row: bool = True,
    autofit_min_width: int = 10,
    autofit_max_width: int = 28,
) -> Path:
    """Convert one CSV file into one formatted XLSX file."""

    LOG.info("========================")
    LOG.info("START convert_csv_to_formatted_xlsx()")
    LOG.info("========================")
    log_path(LOG, "CSV_FILE", csv_path)

    validate_csv_file(csv_path)

    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix(".xlsx")

    log_path(LOG, "XLSX_FILE", xlsx_path)

    rows: list[list[str]] = read_csv_rows(csv_path)

    workbook: Workbook = Workbook()
    worksheet_obj = workbook.active
    if worksheet_obj is None:
        msg = "Workbook did not provide an active worksheet."
        LOG.error(msg)
        raise ValueError(msg)

    worksheet: Worksheet = worksheet_obj
    worksheet.title = safe_sheet_title(csv_path.stem)

    write_rows_to_worksheet(worksheet, rows)
    header_text: str = build_header_text(
        csv_path,
        title_prefix=title_prefix,
    )

    format_worksheet(
        worksheet,
        landscape=landscape,
        repeat_header_row=repeat_header_row,
        autofit_min_width=autofit_min_width,
        autofit_max_width=autofit_max_width,
        header_text=header_text,
    )

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)

    LOG.info(f"Wrote formatted workbook: {xlsx_path}")
    LOG.info("========================")
    LOG.info("END convert_csv_to_formatted_xlsx()")
    LOG.info("========================")

    return xlsx_path


def convert_csv_folder_to_xlsx(
    csv_dir: Path,
    *,
    pattern: str = "*.csv",
    output_dir: Path | None = None,
    title_prefix: str = "Members Signin",
) -> list[Path]:
    """Convert all matching CSV files in a folder to formatted XLSX files."""

    LOG.info("========================")
    LOG.info("START convert_csv_folder_to_xlsx()")
    LOG.info("========================")
    log_path(LOG, "CSV_DIR", csv_dir)

    if not csv_dir.exists():
        msg = f"CSV directory not found: {csv_dir}"
        LOG.error(msg)
        raise FileNotFoundError(msg)

    if not csv_dir.is_dir():
        msg = f"CSV path is not a directory: {csv_dir}"
        LOG.error(msg)
        raise ValueError(msg)

    if output_dir is None:
        output_dir = csv_dir

    output_dir.mkdir(parents=True, exist_ok=True)
    log_path(LOG, "OUTPUT_DIR", output_dir)

    written_paths: list[Path] = []

    for csv_path in sorted(csv_dir.glob(pattern)):
        if not csv_path.is_file():
            continue

        xlsx_path: Path = output_dir / f"{csv_path.stem}.xlsx"
        written_path: Path = convert_csv_to_formatted_xlsx(
            csv_path=csv_path,
            xlsx_path=xlsx_path,
            title_prefix=title_prefix,
        )
        written_paths.append(written_path)

    LOG.info(f"Converted {len(written_paths)} CSV files to XLSX.")
    LOG.info("========================")
    LOG.info("END convert_csv_folder_to_xlsx()")
    LOG.info("========================")

    return written_paths


# === CORE WORKSHEET HELPERS ===


def write_rows_to_worksheet(worksheet: Worksheet, rows: list[list[str]]) -> None:
    """Write raw CSV rows into the worksheet."""

    LOG.info("========================")
    LOG.info("START write_rows_to_worksheet()")
    LOG.info("========================")

    for row in rows:
        worksheet.append(row)

    LOG.info(f"Wrote {len(rows)} rows to worksheet.")
    LOG.info("========================")
    LOG.info("END write_rows_to_worksheet()")
    LOG.info("========================")


def format_worksheet(
    worksheet: Worksheet,
    *,
    landscape: bool,
    repeat_header_row: bool,
    autofit_min_width: int,
    autofit_max_width: int,
    header_text: str,
) -> None:
    """Apply print and style formatting to the worksheet."""

    LOG.info("========================")
    LOG.info("START format_worksheet()")
    LOG.info("========================")

    max_row: int = worksheet.max_row
    max_col: int = worksheet.max_column

    if max_row == 0 or max_col == 0:
        LOG.info("Worksheet is empty. No formatting applied.")
        LOG.info("========================")
        LOG.info("END format_worksheet()")
        LOG.info("========================")
        return

    style_header_row(worksheet)
    style_data_area(worksheet)
    emphasize_record_blocks(worksheet)
    set_column_widths(
        worksheet,
        min_width=autofit_min_width,
        max_width=autofit_max_width,
    )
    set_print_options(
        worksheet,
        landscape=landscape,
        repeat_header_row=repeat_header_row,
        header_text=header_text,
    )

    LOG.info("Applied worksheet formatting successfully.")
    LOG.info("========================")
    LOG.info("END format_worksheet()")
    LOG.info("========================")


def style_header_row(worksheet: Worksheet) -> None:
    """Apply header styling to row 1."""

    LOG.info("========================")
    LOG.info("START style_header_row()")
    LOG.info("========================")

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER

    LOG.info("Styled header row.")
    LOG.info("========================")
    LOG.info("END style_header_row()")
    LOG.info("========================")


def style_data_area(worksheet: Worksheet) -> None:
    """Apply thin borders to all non-header cells."""

    LOG.info("========================")
    LOG.info("START style_data_area()")
    LOG.info("========================")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER

    LOG.info("Styled data area.")
    LOG.info("========================")
    LOG.info("END style_data_area()")
    LOG.info("========================")


def emphasize_record_blocks(worksheet: Worksheet) -> None:
    """Add slightly heavier boxing around each two-row person block.

    Assumes:
    - row 1 is the CSV header
    - each record occupies 2 rows
    - optional spacer rows, if present, are left with thin borders
    """

    LOG.info("========================")
    LOG.info("START emphasize_record_blocks()")
    LOG.info("========================")

    max_col: int = worksheet.max_column
    current_row: int = 2

    while current_row <= worksheet.max_row:
        row_a: int = current_row
        row_b: int = current_row + 1

        if row_b > worksheet.max_row:
            break

        if _is_blank_worksheet_row(worksheet, row_a):
            current_row += 1
            continue

        if _is_blank_worksheet_row(worksheet, row_b):
            current_row += 1
            continue

        for col_idx in range(1, max_col + 1):
            cell_a = worksheet.cell(row=row_a, column=col_idx)
            cell_b = worksheet.cell(row=row_b, column=col_idx)

            left_side = MEDIUM_BLACK_SIDE if col_idx == 1 else THIN_BLACK_SIDE
            right_side = MEDIUM_BLACK_SIDE if col_idx == max_col else THIN_BLACK_SIDE

            cell_a.border = Border(
                left=left_side,
                right=right_side,
                top=MEDIUM_BLACK_SIDE,
                bottom=THIN_BLACK_SIDE,
            )
            cell_b.border = Border(
                left=left_side,
                right=right_side,
                top=THIN_BLACK_SIDE,
                bottom=MEDIUM_BLACK_SIDE,
            )

        current_row += 2

        if current_row <= worksheet.max_row and _is_blank_worksheet_row(
            worksheet, current_row
        ):
            current_row += 1

    LOG.info("Emphasized two-row record blocks.")
    LOG.info("========================")
    LOG.info("END emphasize_record_blocks()")
    LOG.info("========================")


def set_column_widths(
    worksheet: Worksheet,
    *,
    min_width: int,
    max_width: int,
) -> None:
    """Set practical column widths based on content length."""

    LOG.info("========================")
    LOG.info("START set_column_widths()")
    LOG.info("========================")

    for col_idx in range(1, worksheet.max_column + 1):
        column_letter: str = get_column_letter(col_idx)
        max_length: int = 0

        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            text: str = "" if value is None else str(value)
            max_length = max(max_length, len(text))

        adjusted_width: int = max(min_width, min(max_length + 2, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width

    LOG.info("Set column widths.")
    LOG.info("========================")
    LOG.info("END set_column_widths()")
    LOG.info("========================")


def set_print_area_to_used_range(worksheet: Worksheet) -> None:
    """Set the worksheet print area to the used cell range."""

    LOG.info("========================")
    LOG.info("START set_print_area_to_used_range()")
    LOG.info("========================")

    if worksheet.max_row < 1 or worksheet.max_column < 1:
        LOG.info("Worksheet is empty. Print area not set.")
        LOG.info("========================")
        LOG.info("END set_print_area_to_used_range()")
        LOG.info("========================")
        return

    bottom_right: str = f"{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    worksheet.print_area = f"A1:{bottom_right}"

    LOG.info(f"Set print area to A1:{bottom_right}.")
    LOG.info("========================")
    LOG.info("END set_print_area_to_used_range()")
    LOG.info("========================")


def set_page_margins(worksheet: Worksheet) -> None:
    """Apply page margins for printing."""

    LOG.info("========================")
    LOG.info("START set_page_margins()")
    LOG.info("========================")

    worksheet.page_margins.left = DEFAULT_LEFT_MARGIN
    worksheet.page_margins.right = DEFAULT_RIGHT_MARGIN
    worksheet.page_margins.top = DEFAULT_TOP_MARGIN
    worksheet.page_margins.bottom = DEFAULT_BOTTOM_MARGIN
    worksheet.page_margins.header = DEFAULT_HEADER_MARGIN
    worksheet.page_margins.footer = DEFAULT_FOOTER_MARGIN

    LOG.info("Set page margins.")
    LOG.info("========================")
    LOG.info("END set_page_margins()")
    LOG.info("========================")


def set_page_centering(worksheet: Worksheet) -> None:
    """Center the worksheet horizontally on the page."""

    LOG.info("========================")
    LOG.info("START set_page_centering()")
    LOG.info("========================")

    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = False

    LOG.info("Set page centering options.")
    LOG.info("========================")
    LOG.info("END set_page_centering()")
    LOG.info("========================")


def set_header_footer_text(worksheet: Worksheet, *, header_text: str) -> None:
    """Set custom header and footer text for printing."""

    LOG.info("========================")
    LOG.info("START set_header_footer_text()")
    LOG.info("========================")

    worksheet.oddHeader.center.text = header_text
    worksheet.oddFooter.right.text = "Page &[Page] of &[Pages]"

    LOG.info("Set header and footer text.")
    LOG.info("========================")
    LOG.info("END set_header_footer_text()")
    LOG.info("========================")


def build_header_text(
    csv_path: Path,
    *,
    title_prefix: str,
) -> str:
    """Build the printable header text from the CSV filename."""

    LOG.info("========================")
    LOG.info("START build_header_text()")
    LOG.info("========================")

    stem: str = csv_path.stem
    label: str = extract_split_label_from_stem(stem)
    header_text: str = f"{title_prefix} - Last Name {label}"

    LOG.info(f"Built header text: {header_text}")
    LOG.info("========================")
    LOG.info("END build_header_text()")
    LOG.info("========================")

    return header_text


def extract_split_label_from_stem(stem: str) -> str:
    """Extract the final split label from a generated filename stem."""

    parts: list[str] = stem.split("_")
    if parts:
        return parts[-1]

    return stem


def set_print_options(
    worksheet: Worksheet,
    *,
    landscape: bool,
    repeat_header_row: bool,
    header_text: str,
) -> None:
    """Apply print options to the worksheet."""

    LOG.info("========================")
    LOG.info("START set_print_options()")
    LOG.info("========================")

    if landscape:
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE

    worksheet.page_setup.paperSize = LETTER_PAPER_SIZE

    if repeat_header_row:
        worksheet.print_title_rows = "1:1"

    set_print_area_to_used_range(worksheet)
    set_page_margins(worksheet)
    set_page_centering(worksheet)
    set_header_footer_text(worksheet, header_text=header_text)

    worksheet.sheet_view.showGridLines = True

    LOG.info("Applied print options.")
    LOG.info("========================")
    LOG.info("END set_print_options()")
    LOG.info("========================")


# === FILE HELPERS ===


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    """Read all rows from a CSV file."""

    LOG.info("========================")
    LOG.info("START read_csv_rows()")
    LOG.info("========================")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as infile:
        reader = csv.reader(infile)
        rows: list[list[str]] = [list(row) for row in reader]

    LOG.info(f"Read {len(rows)} rows from CSV.")
    LOG.info("========================")
    LOG.info("END read_csv_rows()")
    LOG.info("========================")

    return rows


def validate_csv_file(csv_path: Path) -> None:
    """Validate the CSV file path."""

    if not csv_path.exists():
        msg = f"CSV file not found: {csv_path}"
        LOG.error(msg)
        raise FileNotFoundError(msg)

    if not csv_path.is_file():
        msg = f"CSV path is not a file: {csv_path}"
        LOG.error(msg)
        raise ValueError(msg)

    if csv_path.suffix.lower() != ".csv":
        msg = f"Expected a .csv file: {csv_path}"
        LOG.error(msg)
        raise ValueError(msg)


def safe_sheet_title(raw_title: str) -> str:
    """Return a worksheet title that Excel accepts."""

    invalid_chars: tuple[str, ...] = ("\\", "/", "*", "[", "]", ":", "?")
    clean_title: str = raw_title

    for invalid_char in invalid_chars:
        clean_title = clean_title.replace(invalid_char, "_")

    clean_title = clean_title.strip()

    if not clean_title:
        clean_title = "Sheet1"

    return clean_title[:31]


def _is_blank_worksheet_row(worksheet: Worksheet, row_idx: int) -> bool:
    """Return True when every cell in the worksheet row is blank."""

    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=row_idx, column=col_idx).value
        if value is not None and str(value).strip():
            return False

    return True
